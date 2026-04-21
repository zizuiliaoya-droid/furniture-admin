from django.db import transaction
from openpyxl import load_workbook

from common.file_storage import FileStorageService
from .models import Category, Product, ProductImage


class CategoryService:
    @staticmethod
    def get_tree(dimension: str):
        return Category.objects.filter(
            dimension=dimension, parent__isnull=True
        ).order_by("sort_order", "id")

    @staticmethod
    def get_descendant_ids(category_id: int) -> list[int]:
        ids = [category_id]
        children = Category.objects.filter(parent_id=category_id)
        for child in children:
            ids.extend(CategoryService.get_descendant_ids(child.id))
        return ids

    @staticmethod
    def reorder(items: list[dict]) -> None:
        with transaction.atomic():
            for item in items:
                Category.objects.filter(pk=item["id"]).update(sort_order=item["sort_order"])


class ProductImageService:
    @staticmethod
    def upload(product: Product, files) -> list[ProductImage]:
        images = []
        has_cover = product.images.filter(is_cover=True).exists()

        for i, file in enumerate(files):
            result = FileStorageService.save_image(file, "products")
            image = ProductImage.objects.create(
                product=product,
                image_path=result["path"],
                thumbnail_path=result["thumbnails"].get("small", ""),
                sort_order=product.images.count() + i,
                is_cover=not has_cover and i == 0,
            )
            if not has_cover and i == 0:
                has_cover = True
            images.append(image)
        return images

    @staticmethod
    def delete(image: ProductImage) -> None:
        FileStorageService.delete_file(image.image_path)
        if image.thumbnail_path:
            FileStorageService.delete_file(image.thumbnail_path)
        was_cover = image.is_cover
        product = image.product
        image.delete()
        if was_cover:
            first = product.images.first()
            if first:
                first.is_cover = True
                first.save(update_fields=["is_cover"])

    @staticmethod
    def set_cover(image: ProductImage) -> None:
        ProductImage.objects.filter(product=image.product, is_cover=True).update(is_cover=False)
        image.is_cover = True
        image.save(update_fields=["is_cover"])

    @staticmethod
    def update_order(product: Product, order_data: list[dict]) -> None:
        with transaction.atomic():
            for item in order_data:
                ProductImage.objects.filter(
                    pk=item["id"], product=product
                ).update(sort_order=item["sort_order"])


class ProductImportService:
    @staticmethod
    def import_from_excel(file, user) -> dict:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        success_count = 0
        errors = []

        with transaction.atomic():
            for i, row in enumerate(rows, start=2):
                try:
                    if not row[0]:
                        continue
                    Product.objects.create(
                        name=str(row[0]).strip(),
                        code=str(row[1]).strip() if row[1] else None,
                        description=str(row[2]).strip() if row[2] else "",
                        origin=str(row[3]).strip() if row[3] else "DOMESTIC",
                        min_price=row[4] if row[4] else None,
                        created_by=user,
                    )
                    success_count += 1
                except Exception as e:
                    errors.append({"row": i, "error": str(e)})

        wb.close()
        return {"success": success_count, "errors": errors, "total": len(rows)}
